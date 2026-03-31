"""
Excel reading and writing utilities
"""
from pathlib import Path
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime
from utils.logging_config import get_logger
from config.settings import OUTPUT_DIR

logger = get_logger("ExcelHandler")


def read_excel_urls(file_source) -> tuple[List[str], List[str]]:
    """
    Read URLs from Column B (starting row 3) and headers from Row 2
    Accepts a file path (str/Path) or a file-like object (BytesIO)
    Returns: (urls, headers)
    """
    wb = openpyxl.load_workbook(file_source)
    ws = wb.active
    
    # Extract headers from row 2 (starting from column C)
    headers = []
    for col_idx in range(3, ws.max_column + 1):  # C onwards
        cell = ws.cell(row=2, column=col_idx)
        if cell.value:
            headers.append(str(cell.value).strip())
    
    # Extract URLs from column B (starting row 3)
    urls = []
    for row_idx in range(3, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=2)  # Column B
        if cell.value:
            url = str(cell.value).strip()
            if url.startswith("http"):
                urls.append(url)
    
    wb.close()
    
    logger.info(f"Extracted {len(urls)} URLs and {len(headers)} column headers")
    return urls, headers


def write_excel_output(results: List[Dict[str, Any]], headers: List[str]) -> str:
    """
    Write audit results to Excel
    results: List of {url, summaries: {column: text}}
    headers: List of column names
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SEO Audit Report"
    
    # Header row
    ws.cell(row=1, column=1, value="URL")
    for col_idx, header in enumerate(headers, start=2):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Data rows
    for row_idx, result in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1, value=result["url"])
        
        summaries = result.get("summaries", {})
        for col_idx, header in enumerate(headers, start=2):
            summary = summaries.get(header, "No data")
            cell = ws.cell(row=row_idx, column=col_idx, value=summary)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 50
    for col_idx in range(2, len(headers) + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 40
    
    # Save file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = OUTPUT_DIR / f"seo_audit_{timestamp}.xlsx"
    wb.save(output_path)
    wb.close()
    
    logger.info(f"Excel report saved: {output_path}")
    return str(output_path)